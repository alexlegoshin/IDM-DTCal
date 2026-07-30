"""
Расчёт ожидаемого напряжения/погрешности и запись результата в Excel.

Заменяет связку CSV + отдельная команда analyze из IVTrace: DTCal сразу
после измерения знает модель датчика (и её I_ном), поэтому считает
погрешность на месте и пишет один .xlsx.

Структура книги (три листа) выбрана так, чтобы измеренное было отделено от
вычисленного:

    «Инфо»        — метаданные сессии (что, чем и с какими параметрами снимали);
    «Данные»      — ТОЛЬКО первичные измерения: ток задания и напряжение с датчика;
    «Погрешность» — производный расчёт: методика, дисклеймер и таблица
                    U ожид./γ, посчитанные программой по номинальной
                    характеристике.

Разделение принципиальное: номинальная характеристика задаётся оператором и
может не совпадать с ТЗ/ТУ на конкретную партию датчиков, поэтому лист
«Погрешность» несёт явное предупреждение и ссылку на первичные столбцы.

Оформление листов делается здесь же (ширина колонок, форматы чисел,
закреплённая шапка), чтобы готовый файл не приходилось подгонять руками.
Погрешность выводится с явным знаком — направление отклонения от
номинальной характеристики важно для калибровки.
"""
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sensors import (
    DEFAULT_SCALE,
    ERROR_FORMULA_TEXT,
    EXPECTED_FORMULA_TEXT,
    NORMALIZATION_TEXT,
    OutputScale,
    error_percent,
    expected_voltage,
    scale_from_params,
)


# Внутренние (англоязычные) имена колонок -> заголовки для Excel.
# В DataFrame имена остаются латиницей: по ним работают тесты и код GUI,
# а кириллица нужна только в готовом отчёте для оператора.
COLUMN_TITLES = {
    'Timestamp': 'Время',
    'Branch': 'Направление',
    'I_set_A': 'I задан., А',
    'V_meas_V': 'U измер., В',
    'V_expected_V': 'U ожид., В',
    'Error_percent': 'Погрешность, %',
}

# Что считается первичным измерением, а что — расчётом. Столбцы листа
# «Данные» приходят прямо с приборов; всё остальное вычислено программой.
RAW_COLUMNS = ['Timestamp', 'Branch', 'I_set_A', 'V_meas_V']
ERROR_SHEET_COLUMNS = ['Branch', 'I_set_A', 'V_meas_V', 'V_expected_V', 'Error_percent']

MEASURED_COLUMNS_HINT = "«I задан., А» и «U измер., В» на листе «Данные»"

DISCLAIMER_TEXT = (
    "ВНИМАНИЕ. Погрешность здесь — расчётная величина, а не результат измерения. "
    "Она получена по номинальной характеристике и способу нормирования, указанным выше, "
    "и может быть посчитана некорректно, если эти исходные данные не совпадают с ТЗ и ТУ "
    "на конкретный датчик. Перед использованием отчёта сверьтесь с ТЗ и ТУ: точки выходной "
    "характеристики, номинальный ток и нормирующее значение. "
    f"Опирайтесь на реально измеренные данные — это только столбцы {MEASURED_COLUMNS_HINT}; "
    "столбцы «U ожид., В» и «Погрешность, %» вычислены программой из них."
)

BRANCH_TITLES = {'forward': 'прямое', 'reverse': 'обратное'}

# Формат с явным знаком: секции Excel — положительные;отрицательные;ноль.
SIGNED_PERCENT_FORMAT = '+0.000;-0.000;0.000'

NUMBER_FORMATS = {
    'I_set_A': '0.00',
    'V_meas_V': '0.0000',
    'V_expected_V': '0.0000',
    'Error_percent': SIGNED_PERCENT_FORMAT,
}

_HEADER_FILL = PatternFill('solid', fgColor='EFEFEF')
_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11)
_WARN_FONT = Font(bold=True, color='B00020')
_WARN_FILL = PatternFill('solid', fgColor='FDECEA')


def build_report(df: pd.DataFrame, i_nom: float, scale: OutputScale = DEFAULT_SCALE) -> pd.DataFrame:
    """
    Добавляет к сырым данным измерения колонки V_expected_V и Error_percent (со знаком).

    scale — точки номинальной выходной характеристики датчика; по умолчанию
    симметричные ±4 В (см. sensors.DEFAULT_SCALE).

    Пустой DataFrame (измерение остановили до первой точки) не является
    ошибкой: возвращаем пустую таблицу с полным набором колонок, иначе
    обращение к df['I_set_A'] упало бы с KeyError.
    """
    if i_nom is None or i_nom <= 0:
        raise ValueError(f"Номинальный ток должен быть положительным, получено {i_nom!r}")

    df = df.copy()

    if df.empty:
        for col in list(COLUMN_TITLES):
            if col not in df.columns:
                df[col] = pd.Series(dtype='float64')
        return df

    df['V_expected_V'] = df['I_set_A'].apply(lambda i: expected_voltage(i, i_nom, scale))
    df['Error_percent'] = df.apply(
        lambda r: error_percent(r['V_meas_V'], r['V_expected_V'], scale.span), axis=1
    )
    return df


def _autosize(worksheet, df_like_widths: dict, min_width: int = 10, max_width: int = 22):
    for idx, width in df_like_widths.items():
        letter = get_column_letter(idx)
        worksheet.column_dimensions[letter].width = max(min_width, min(max_width, width))


def _style_header(worksheet, ncols: int, row: int = 1):
    for col in range(1, ncols + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.freeze_panes = worksheet.cell(row=row + 1, column=1).coordinate


def _format_table(worksheet, columns, header_row: int = 1):
    """Числовые форматы по смыслу колонки + ширины. Общее для «Данных» и «Погрешности»."""
    _style_header(worksheet, len(columns), row=header_row)

    widths = {}
    for idx, name in enumerate(columns, start=1):
        widths[idx] = len(COLUMN_TITLES.get(name, name)) + 4
        fmt = NUMBER_FORMATS.get(name)
        if fmt is None:
            continue
        for row in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=idx)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal='right')

    if 'Timestamp' in columns:
        widths[columns.index('Timestamp') + 1] = 20
    _autosize(worksheet, widths)


def _format_info_sheet(worksheet):
    _style_header(worksheet, 2)
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1).font = Font(bold=True)
    worksheet.column_dimensions['A'].width = 38
    worksheet.column_dimensions['B'].width = 46
    for row in range(1, worksheet.max_row + 1):
        worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='left', wrap_text=False)


def _wrapped_height(text: str, width_chars: float, line_height: float = 14.5) -> float:
    """
    Оценка высоты ячейки с переносом по словам, пункты.

    openpyxl не умеет автоподбор высоты при wrap_text, а без явной высоты
    Excel показывает только первую строку — методика и предупреждение
    оказались бы обрезанными. Оценка грубая (моноширинное приближение), но
    ошибается в большую сторону, что безопасно.
    """
    lines = 0
    for paragraph in str(text).split("\n"):
        lines += max(1, -(-len(paragraph) // max(int(width_chars), 1)))
    return max(line_height, lines * line_height)


def _method_rows(i_nom: float, scale: OutputScale) -> list:
    """Пары «параметр — значение», описывающие способ расчёта погрешности."""
    return [
        ("Номинальный ток I ном., А", f"{i_nom:g}"),
        ("Точки номинальной характеристики", scale.describe_points()),
        ("Ожидаемое напряжение U ожид.", EXPECTED_FORMULA_TEXT),
        ("Формула погрешности", ERROR_FORMULA_TEXT),
        ("Нормирующее значение U норм., В",
         f"{scale.span:.4f}   (= V(+I ном.) − V(−I ном.), размах шкалы)"),
        ("Нормировка", NORMALIZATION_TEXT),
        ("Знак погрешности", "«+» выход выше номинальной характеристики, «−» ниже"),
    ]


def _write_error_sheet(writer, df: pd.DataFrame, i_nom: float, scale: OutputScale) -> None:
    """
    Пишет лист «Погрешность»: блок методики, дисклеймер и таблицу расчёта.

    Таблица начинается ниже текстового блока, поэтому startrow вычисляется, а
    не задан константой — иначе любая правка методики молча разъехалась бы с
    форматированием.
    """
    method = _method_rows(i_nom, scale)

    # Раскладка (1-based): заголовок, пусто, «Как считается», строки методики,
    # пусто, дисклеймер на 3 строки, пусто, шапка таблицы.
    row_title = 1
    row_section = 3
    row_method_start = 4
    row_method_end = row_method_start + len(method) - 1
    row_warn = row_method_end + 2
    warn_height_rows = 3
    header_row = row_warn + warn_height_rows + 1

    out = df.reindex(columns=[c for c in ERROR_SHEET_COLUMNS if c in df.columns]).copy()
    columns = list(out.columns)
    if 'Branch' in out.columns:
        out['Branch'] = out['Branch'].map(lambda b: BRANCH_TITLES.get(b, b))
    out = out.rename(columns=COLUMN_TITLES)
    # startrow 0-based, header_row 1-based — шапку пишет сам to_excel.
    out.to_excel(writer, sheet_name='Погрешность', index=False, startrow=header_row - 1)

    ws = writer.sheets['Погрешность']
    ncols = max(len(columns), 2)

    ws.cell(row=row_title, column=1, value="Расчёт погрешности").font = _TITLE_FONT
    ws.cell(row=row_section, column=1, value="Как считается").font = _SECTION_FONT

    for offset, (name, value) in enumerate(method):
        r = row_method_start + offset
        key_cell = ws.cell(row=r, column=1, value=name)
        key_cell.font = Font(bold=True)
        key_cell.alignment = Alignment(vertical='top')
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.merge_cells(start_row=row_warn, start_column=1,
                   end_row=row_warn + warn_height_rows - 1, end_column=ncols)
    warn = ws.cell(row=row_warn, column=1, value=DISCLAIMER_TEXT)
    warn.font = _WARN_FONT
    warn.fill = _WARN_FILL
    warn.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    _format_table(ws, columns, header_row=header_row)
    # Ширина колонок A/B задана под текст методики, а не под данные таблицы.
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 46

    # Высоты считаем последними: нужны итоговые ширины колонок.
    for offset, (_, value) in enumerate(method):
        ws.row_dimensions[row_method_start + offset].height = _wrapped_height(value, 46)

    warn_width = sum(
        ws.column_dimensions[get_column_letter(c)].width or 10 for c in range(1, ncols + 1)
    )
    warn_total = _wrapped_height(DISCLAIMER_TEXT, warn_width)
    for r in range(row_warn, row_warn + warn_height_rows):
        ws.row_dimensions[r].height = warn_total / warn_height_rows


def write_report_xlsx(xlsx_path: Path, df: pd.DataFrame, params: dict) -> None:
    """
    Пишет .xlsx с тремя листами:
        «Инфо»        — метаданные сессии;
        «Данные»      — первичные измерения (ток задания + напряжение датчика);
        «Погрешность» — методика, дисклеймер и расчёт U ожид./γ.

    Точки характеристики берутся из params (V_minus/V_zero/V_plus); если их
    там нет — используются значения по умолчанию ±4 В.
    """
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    scale = scale_from_params(params)
    i_nom = params['i_nom']

    finite_errors = df['Error_percent'].dropna() if 'Error_percent' in df.columns else pd.Series(dtype='float64')

    info_rows = [
        ("Датчик", params['model']),
        ("Номинальный ток, А", i_nom),
        ("Выходная характеристика (номинал)", scale.describe_points()),
        ("Диапазон возбуждения, А", f"{params['I_start']}..{params['I_stop']}, шаг {params['I_step']}"),
        ("Ограничение напряжения источника, В", params['V_limit']),
        ("Обе полярности", "сняты автоматически через плату реле (см. колонку «Направление»)"),
        ("Задержка установки, с", params['delay']),
        ("Задержка охлаждения, с", params['cooling_delay']),
        # Пустая строка в xlsx читается обратно как NaN — пишем прочерк.
        ("Комментарий", params.get('label') or "—"),
        ("Время измерения", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("Всего точек", len(df)),
        ("Реально измеренные данные", MEASURED_COLUMNS_HINT),
        ("Расчёт погрешности", "см. лист «Погрешность» — там же методика и предупреждение"),
        ("Нормирующее значение погрешности, В", scale.span),
        ("Макс. |погрешность|, %", round(float(finite_errors.abs().max()), 4) if not finite_errors.empty else "—"),
        ("Средняя погрешность (со знаком), %", round(float(finite_errors.mean()), 4) if not finite_errors.empty else "—"),
    ]
    info_df = pd.DataFrame(info_rows, columns=["Параметр", "Значение"])

    raw_columns = [c for c in RAW_COLUMNS if c in df.columns]
    raw = df.reindex(columns=raw_columns).copy()
    if 'Branch' in raw.columns:
        raw['Branch'] = raw['Branch'].map(lambda b: BRANCH_TITLES.get(b, b))
    raw = raw.rename(columns=COLUMN_TITLES)

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        info_df.to_excel(writer, sheet_name='Инфо', index=False)
        raw.to_excel(writer, sheet_name='Данные', index=False)
        _write_error_sheet(writer, df, i_nom, scale)

        _format_info_sheet(writer.sheets['Инфо'])
        _format_table(writer.sheets['Данные'], raw_columns)
