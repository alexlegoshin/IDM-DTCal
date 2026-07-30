import pandas as pd
import pytest
from openpyxl import load_workbook

from report import build_report, write_report_xlsx, SIGNED_PERCENT_FORMAT
from sensors import OutputScale, scale_from_params


def _params(**over):
    params = {
        'model': 'ДТ100А1',
        'i_nom': 100.0,
        'label': 'OrchSensor',
        'I_start': 0.0, 'I_stop': 10.0, 'I_step': 5.0,
        'V_limit': 5.0, 'delay': 0.1, 'cooling_delay': 0.2,
    }
    params.update(over)
    return params


def test_build_report_adds_expected_and_error_columns():
    # Номинал (симметричная шкала ±4 В), i_nom=100: V_ожид(I) = 4.0 * I / 100.
    # I=0 -> 0 В, I=+100 -> +4 В, I=-100 -> -4 В.
    df = pd.DataFrame([
        {'Timestamp': '2026-01-01T00:00:00', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.0},
        {'Timestamp': '2026-01-01T00:00:00', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.0},
        {'Timestamp': '2026-01-01T00:00:00', 'Branch': 'reverse', 'I_set_A': -50.0, 'V_meas_V': -2.0},
    ])

    result = build_report(df, i_nom=100.0)

    assert 'V_expected_V' in result.columns
    assert 'Error_percent' in result.columns
    assert len(result) == 3
    assert list(result['V_expected_V']) == pytest.approx([0.0, 2.0, -2.0])
    # Измеренное совпало с ожидаемым во всех точках -> погрешность 0.
    assert list(result['Error_percent']) == pytest.approx([0.0, 0.0, 0.0], abs=1e-9)


def test_build_report_error_has_sign():
    """Погрешность сохраняет знак: «+» если измерено выше ожидаемого."""
    # i_nom=100: V_ожид(50) = +2.0 В
    df = pd.DataFrame([
        {'Timestamp': '2026-01-01T00:00:00', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.1},
        {'Timestamp': '2026-01-01T00:00:00', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 1.9},
    ])

    result = build_report(df, i_nom=100.0)

    # γ = (U изм. - U ожид.) / U норм. * 100, U норм. = 8 В
    # (2.1 - 2.0) / 8 * 100 = +1.25 % ; (1.9 - 2.0) / 8 * 100 = -1.25 %
    assert result.iloc[0]['Error_percent'] == pytest.approx(1.25)
    assert result.iloc[1]['Error_percent'] == pytest.approx(-1.25)


def test_build_report_reverse_branch_is_negative_voltage():
    """
    Регрессия: раньше шкала считалась однополярной (2..10 В со «смещённым
    нулём»). Реальный выход биполярный, и на обратной ветви ожидаемое
    напряжение обязано быть отрицательным.
    """
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'reverse', 'I_set_A': -100.0, 'V_meas_V': -4.0},
    ])

    result = build_report(df, i_nom=100.0)

    assert result.iloc[0]['V_expected_V'] == pytest.approx(-4.0)
    assert result.iloc[0]['Error_percent'] == pytest.approx(0.0, abs=1e-9)


def test_build_report_honours_shifted_zero_scale():
    """
    Смещённый ноль (случай ДТ500А1): ветви имеют разный наклон, и датчик,
    точно попадающий в свою характеристику, должен давать нулевую погрешность.
    """
    scale = OutputScale(v_minus=-3.96, v_zero=0.04, v_plus=4.04)
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.04},
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 500.0, 'V_meas_V': 4.04},
        {'Timestamp': 'x', 'Branch': 'reverse', 'I_set_A': -500.0, 'V_meas_V': -3.96},
    ])

    result = build_report(df, i_nom=500.0, scale=scale)

    assert list(result['Error_percent']) == pytest.approx([0.0, 0.0, 0.0], abs=1e-9)
    # Размах прежний (8 В), поэтому нормировка не поехала.
    assert scale.span == pytest.approx(8.0)


def test_build_report_on_empty_dataframe_returns_columns_not_crash():
    """
    Регрессия: если измерение остановили до первой точки, DataFrame пуст и
    обращение к df['I_set_A'] падало с KeyError, унося с собой весь сеанс.
    """
    result = build_report(pd.DataFrame(), i_nom=100.0)

    assert result.empty
    for col in ('I_set_A', 'V_meas_V', 'V_expected_V', 'Error_percent'):
        assert col in result.columns


def test_build_report_rejects_nonpositive_i_nom():
    """i_nom=0 дал бы деление на ноль в expected_voltage()."""
    df = pd.DataFrame([{'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.0}])

    with pytest.raises(ValueError):
        build_report(df, i_nom=0.0)


def test_build_report_keeps_nan_readings_as_nan():
    """Неудачное чтение (NaN) не должно превратиться в погрешность-число."""
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': float('nan')},
    ])

    result = build_report(df, i_nom=100.0)

    assert pd.isna(result.iloc[0]['Error_percent'])


def test_write_report_xlsx_creates_file_with_three_sheets(tmp_path):
    df = pd.DataFrame([
        {'Timestamp': '2026-01-01T00:00:00', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.0},
    ])
    params = _params()
    xlsx_path = tmp_path / "test_report.xlsx"

    df_with_error = build_report(df, i_nom=params['i_nom'])
    write_report_xlsx(xlsx_path, df_with_error, params)

    assert xlsx_path.exists()

    xl = pd.ExcelFile(xlsx_path)
    assert 'Инфо' in xl.sheet_names
    assert 'Данные' in xl.sheet_names
    assert 'Погрешность' in xl.sheet_names

    data_df = pd.read_excel(xlsx_path, sheet_name='Данные')
    assert len(data_df) == 1
    # В отчёте заголовки человекочитаемые (внутри DataFrame остаются латиницей).
    assert 'I задан., А' in data_df.columns
    assert 'U измер., В' in data_df.columns


def test_data_sheet_holds_only_measured_columns(tmp_path):
    """
    Первичные измерения отделены от расчёта: на листе «Данные» не должно
    быть вычисленных столбцов, иначе их легко принять за показания прибора.
    """
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.1},
    ])
    xlsx_path = tmp_path / "split.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    data_df = pd.read_excel(xlsx_path, sheet_name='Данные')
    assert 'U ожид., В' not in data_df.columns
    assert 'Погрешность, %' not in data_df.columns


def test_error_sheet_states_method_and_disclaimer(tmp_path):
    """
    Лист «Погрешность» обязан объяснять, как считалось, и предупреждать, что
    расчёт может не совпасть с ТЗ/ТУ — иначе цифру примут за истину.
    """
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.1},
    ])
    xlsx_path = tmp_path / "method.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    ws = load_workbook(xlsx_path)['Погрешность']
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )

    assert "Расчёт погрешности" in text
    assert "Нормирующее значение" in text
    assert "ГОСТ 8.401-80" in text
    # Использованные точки характеристики видны прямо в файле.
    assert "V(+I ном.)" in text
    assert "+4.0000" in text
    # Дисклеймер и указание на первичные столбцы.
    assert "ТЗ" in text and "ТУ" in text
    assert "U измер., В" in text
    # Сами вычисленные столбцы тоже на этом листе.
    assert "U ожид., В" in text
    assert "Погрешность, %" in text


def test_error_sheet_table_is_formatted(tmp_path):
    """Шапка таблицы расчёта закреплена и погрешность выводится со знаком."""
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.1},
    ])
    xlsx_path = tmp_path / "fmt.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    ws = load_workbook(xlsx_path)['Погрешность']

    header_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == 'Направление'
    )
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    err_col = headers.index('Погрешность, %') + 1

    assert ws.cell(row=header_row, column=1).font.bold is True
    assert ws.freeze_panes == ws.cell(row=header_row + 1, column=1).coordinate
    assert ws.cell(row=header_row + 1, column=err_col).number_format == SIGNED_PERCENT_FORMAT
    assert ws.cell(row=header_row + 1, column=err_col).value == pytest.approx(1.25)


def test_error_sheet_shows_shifted_scale_actually_used(tmp_path):
    """В отчёт попадают те точки характеристики, по которым реально считали."""
    df = pd.DataFrame([{'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.04}])
    params = _params(model='ДТ500А1', i_nom=500.0, V_minus=-3.96, V_zero=0.04, V_plus=4.04)
    xlsx_path = tmp_path / "shifted.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, 500.0, scale_from_params(params)), params)

    ws = load_workbook(xlsx_path)['Погрешность']
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )

    assert "-3.9600" in text
    assert "+0.0400" in text
    assert "+4.0400" in text


def test_write_report_xlsx_translates_branch_names(tmp_path):
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.0},
        {'Timestamp': 'x', 'Branch': 'reverse', 'I_set_A': -50.0, 'V_meas_V': -2.0},
    ])
    xlsx_path = tmp_path / "branches.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    data_df = pd.read_excel(xlsx_path, sheet_name='Данные')
    assert list(data_df['Направление']) == ['прямое', 'обратное']


def test_write_report_xlsx_formats_data_sheet_for_reading(tmp_path):
    """
    Оформление должно быть готовым: закреплённая жирная шапка и заданные
    ширины — чтобы файл не подгоняли руками.
    """
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.1},
    ])
    xlsx_path = tmp_path / "formatted.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    ws = load_workbook(xlsx_path)['Данные']
    assert ws.freeze_panes == 'A2'
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.column_dimensions['A'].width > 0


def test_write_report_xlsx_info_sheet_contains_error_summary(tmp_path):
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 50.0, 'V_meas_V': 2.1},
        {'Timestamp': 'x', 'Branch': 'reverse', 'I_set_A': -50.0, 'V_meas_V': -2.1},
    ])
    xlsx_path = tmp_path / "info.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    info_df = pd.read_excel(xlsx_path, sheet_name='Инфо')
    keys = list(info_df['Параметр'])
    assert "Датчик" in keys
    assert "Нормирующее значение погрешности, В" in keys
    assert "Выходная характеристика (номинал)" in keys
    assert any("Макс" in k for k in keys)


def test_write_report_xlsx_creates_missing_directory(tmp_path):
    """Оператор мог выбрать папку, которой уже нет — не падаем."""
    df = pd.DataFrame([{'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': 0.0}])
    xlsx_path = tmp_path / "nested" / "deeper" / "r.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    assert xlsx_path.exists()


def test_write_report_xlsx_handles_all_nan_data(tmp_path):
    """Полностью неудачное измерение всё равно должно записаться (со сводкой «—»)."""
    df = pd.DataFrame([
        {'Timestamp': 'x', 'Branch': 'forward', 'I_set_A': 0.0, 'V_meas_V': float('nan')},
    ])
    xlsx_path = tmp_path / "nan.xlsx"

    write_report_xlsx(xlsx_path, build_report(df, i_nom=100.0), _params())

    assert xlsx_path.exists()
